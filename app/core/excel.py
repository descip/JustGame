from __future__ import annotations

from io import BytesIO
from typing import Iterable, Sequence, Any, cast

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


def _autosize_columns(ws: Worksheet) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 45)


def make_workbook(title: str, headers: list[str], rows: Iterable[Sequence[Any]]) -> bytes:
    """
    Создаёт Excel-файл (xlsx) в памяти и возвращает bytes.
    rows: любая последовательность значений (int/float/str/None и т.д.)
    """
    wb = Workbook()
    ws = cast(Worksheet, wb.active)   # для Pylance: active существует и это Worksheet
    ws.title = title[:31]            # Excel ограничение на имя листа

    ws.append(headers)
    for r in rows:
        ws.append(list(r))

    _autosize_columns(ws)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
